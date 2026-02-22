from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from .analytics import compute_analytics


def _autosize_columns(ws) -> None:
    """Автопідбір ширини колонок по максимальній довжині значень"""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)


def _fmt_float(x: Optional[float]) -> Optional[float]:
    """Безпечне перетворення в float (інакше None)"""
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None


def build_excel(
    *,
    query: str, # пошуковий запит
    items: List[Dict[str, Any]], # нормалізовані товари
    total: int | None = None, # total з API
    limit: int | None = None, # limit з API
    offset: int | None = None, # offset з API
    sort: str | None = None, # сортування
) -> bytes:
    """Excel експорт: Items + Analytics + Charts"""
    wb = Workbook()

    header_font = Font(bold=True)  # стиль заголовків
    # Items (таблиця товарів)
    ws = wb.active
    ws.title = "Items"

    headers = [
        "#",
        "Title",
        "Category",
        "Condition",
        "Price Value",
        "Price Currency",
        "Shipping Value",
        "Shipping Currency",
        "Total (Price+Ship)",
        "Currency",
        "Country",
        "Seller feedback",
        "Item ID",
        "Web URL",
    ]
    ws.append(headers)

    # формат заголовків
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # запис рядків товарів
    for idx, it in enumerate(items or [], start=1):
        pv = _fmt_float(it.get("price_value"))
        sv = _fmt_float(it.get("shipping_value"))
        cur = (it.get("price_currency") or it.get("shipping_currency") or "") or None # валюта
        total_val = (pv or 0.0) + (sv or 0.0) if (pv is not None or sv is not None) else None # total

        ws.append(
            [
                idx,
                it.get("title"),
                it.get("category"),
                it.get("condition"),
                pv,
                it.get("price_currency"),
                sv,
                it.get("shipping_currency"),
                total_val,
                cur,
                it.get("location_country"),
                _fmt_float(it.get("seller_feedback")),
                it.get("itemId"),
                it.get("web_url"),
            ]
        )

    # formats для числових колонок + вирівнювання
    for row in ws.iter_rows(min_row=2):
        row[4].number_format = "0.00" # price
        row[6].number_format = "0.00" # shipping
        row[8].number_format = "0.00" # total
        row[11].number_format = "0" # seller feedback (ціле)

        for c in row:
            if c.column in (1, 5, 7, 9, 12): # де зручно праве вирівнювання
                c.alignment = Alignment(vertical="top", horizontal="right", wrap_text=True)
            else:
                c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    _autosize_columns(ws) # підбір ширини

    # Analytics (таблиця метрик/статистик)
    ws2 = wb.create_sheet("Analytics")

    analytics = compute_analytics(items or [])  # обчислення аналітики
    cur = analytics.get("currency_most_common") or ""  # найчастіша валюта

    ws2.append(["Metric", "Value"])
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font

    # метадані експорту
    meta_rows = [
        ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Query", query),
        ("Sort", sort or ""),
        ("API total", total),
        ("Limit", limit),
        ("Offset", offset),
        ("Items exported", len(items or [])),
        ("Most common currency", cur),
    ]
    for k, v in meta_rows:
        ws2.append([k, v])

    ws2.append([])

    # секції статистик
    ws2.append(["Section", "Metric", "Value"])
    for c in ws2[ws2.max_row]:
        c.font = header_font

    def add_section(section: str, block: Dict[str, Any]):
        """Додає stats-блок у вигляді рядків (section, metric, value)"""
        for metric, val in (block or {}).items():
            ws2.append([section, metric, val])

    add_section("price", analytics.get("price"))
    add_section("shipping", analytics.get("shipping"))
    add_section("total", analytics.get("total"))
    add_section("seller_feedback", analytics.get("seller_feedback"))

    ws2.append([])

    # пропуски по полях
    ws2.append(["Missing fields", "field", "count"])
    for c in ws2[ws2.max_row]:
        c.font = header_font
    for k, v in (analytics.get("missing") or {}).items():
        ws2.append(["", k, v])

    ws2.append([])

    # топи (країни/категорії/стани)
    ws2.append(["Top (by count)", "Key", "Count"])
    for c in ws2[ws2.max_row]:
        c.font = header_font

    top_block = analytics.get("top") or {}
    for group_name, entries in top_block.items():
        ws2.append([group_name, "", ""])  # назва групи
        for e in entries or []:
            ws2.append(["", e.get("key"), e.get("count")])

    ws2.freeze_panes = "A2"
    _autosize_columns(ws2)

    # Charts (дані для графіків + діаграми)
    ws3 = wb.create_sheet("Charts")

    ws3.append(["Top Countries", "", ""])
    ws3["A1"].font = header_font

    # таблиця для top countries
    ws3.append(["Country", "Count"])
    ws3["A2"].font = header_font
    ws3["B2"].font = header_font

    countries = (analytics.get("top", {}).get("countries") or [])[:7]  # топ-7 країн
    for e in countries:
        ws3.append([e.get("key"), e.get("count")])

    # Bar chart для країн
    if countries:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Top Countries (count)"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Country"

        data = Reference(ws3, min_col=2, min_row=2, max_row=2 + len(countries))
        cats = Reference(ws3, min_col=1, min_row=3, max_row=2 + len(countries))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()  # показ значень на барах
        chart.dataLabels.showVal = True
        ws3.add_chart(chart, "D2")  # позиція графіка

    ws3.append([])
    start_row = ws3.max_row + 1  # рядок, з якого починається histogram
    ws3.append(["Total histogram", "", ""])
    ws3[f"A{start_row}"].font = header_font

    # таблиця бінів гістограми
    ws3.append(["From", "To", "Count"])
    ws3[f"A{start_row+1}"].font = header_font
    ws3[f"B{start_row+1}"].font = header_font
    ws3[f"C{start_row+1}"].font = header_font

    bins = (analytics.get("hist", {}).get("total", {}).get("bins") or [])  # біни total
    for b in bins:
        ws3.append([b.get("from"), b.get("to"), b.get("count")])

    # Histogram chart
    if bins:
        label_col = 4  # допоміжні колонки для label+count
        ws3.cell(row=start_row + 1, column=label_col, value="Bin").font = header_font
        ws3.cell(row=start_row + 1, column=label_col + 1, value="Count").font = header_font

        # формування текстового label "a-b"
        for i, b in enumerate(bins, start=1):
            r = start_row + 1 + i
            ws3.cell(row=r, column=label_col, value=f"{b['from']:.2f}-{b['to']:.2f}")
            ws3.cell(row=r, column=label_col + 1, value=b["count"])

        hist_chart = BarChart()
        hist_chart.type = "col"
        hist_chart.title = "Total price histogram (count)"
        hist_chart.y_axis.title = "Count"
        hist_chart.x_axis.title = "Bin"

        data = Reference(ws3, min_col=label_col + 1, min_row=start_row + 1, max_row=start_row + 1 + len(bins))
        cats = Reference(ws3, min_col=label_col, min_row=start_row + 2, max_row=start_row + 1 + len(bins))
        hist_chart.add_data(data, titles_from_data=True)
        hist_chart.set_categories(cats)
        ws3.add_chart(hist_chart, f"D{start_row+2}") # позиція графіка

    _autosize_columns(ws3) # підбір ширин для Charts

    bio = BytesIO() # запис xlsx в пам’ять
    wb.save(bio)
    return bio.getvalue() # bytes для StreamingResponse
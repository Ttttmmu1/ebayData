from __future__ import annotations

from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference


def _safe_sheet_title(title: str) -> str:
    t = (title or "").strip()[:31]
    return t or "Sheet"


def _set_col_widths(ws, columns: List[str], rows: List[Dict[str, Any]], max_width: int = 45) -> None:
    # simple width estimator
    widths = {c: min(max(len(str(c)), 10), max_width) for c in columns}
    for r in rows[:200]:
        for c in columns:
            v = r.get(c, "")
            widths[c] = min(max(widths[c], len(str(v)) if v is not None else 0), max_width)

    for i, c in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[c] + 2


def _write_table(ws, columns: List[str], rows: List[Dict[str, Any]], title: Optional[str] = None) -> None:
    ws.freeze_panes = "A2"

    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
        ws.append([])  # spacer

    # header
    ws.append(columns)
    header_row = ws.max_row
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # body
    for r in rows:
        ws.append([r.get(c, "") for c in columns])

    # autofilter
    first_data_row = header_row
    last_row = ws.max_row
    last_col = get_column_letter(len(columns)) if columns else "A"
    ws.auto_filter.ref = f"A{first_data_row}:{last_col}{last_row}"

    _set_col_widths(ws, columns, rows)


def _write_kv(ws, kv: List[List[Any]], start_row: int = 1, start_col: int = 1, title: Optional[str] = None) -> int:
    r = start_row
    c = start_col

    if title:
        ws.cell(row=r, column=c, value=title).font = Font(bold=True, size=14)
        r += 2

    for k, v in kv:
        ws.cell(row=r, column=c, value=str(k)).font = Font(bold=True)
        ws.cell(row=r, column=c + 1, value=v)
        r += 1

    ws.column_dimensions[get_column_letter(c)].width = 28
    ws.column_dimensions[get_column_letter(c + 1)].width = 60
    return r + 1


def build_filtered_excel(
    *,
    dataset_name: str,
    mode_text: str,
    columns: List[str],
    rows: List[Dict[str, Any]],
    filter_col: str,
    filter_text: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title("Data")

    title = f"Filtered export: {dataset_name}"
    _write_table(ws, columns, rows, title=title)

    meta = wb.create_sheet(_safe_sheet_title("Meta"))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _write_kv(
        meta,
        [
            ["Dataset", dataset_name],
            ["Mode", mode_text],
            ["Generated", now],
            ["Filter column", filter_col or "(all columns)"],
            ["Filter text", filter_text or ""],
            ["Rows exported", len(rows)],
            ["Columns", len(columns)],
        ],
        title="Export metadata",
    )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_report_excel(
    *,
    dataset_name: str,
    mode_text: str,
    columns: List[str],
    rows: List[Dict[str, Any]],
    filter_col: str,
    filter_text: str,
    numeric_col: str,
    colstats: Dict[str, Any],
    hist_labels: List[str],
    hist_counts: List[int],
    cat_col: str,
    top_labels: List[str],
    top_counts: List[int],
) -> bytes:
    wb = Workbook()

    # 1) Data sheet
    ws_data = wb.active
    ws_data.title = _safe_sheet_title("Data")
    _write_table(ws_data, columns, rows, title=f"Dataset report: {dataset_name}")

    # 2) Analytics sheet
    ws_a = wb.create_sheet(_safe_sheet_title("Analytics"))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    st = (colstats or {}).get("stats") or {}
    parse_ratio = (colstats or {}).get("parse_ratio", 0.0)

    kv1 = [
        ["Dataset", dataset_name],
        ["Mode", mode_text],
        ["Generated", now],
        ["Filter column", filter_col or "(all columns)"],
        ["Filter text", filter_text or ""],
        ["Rows in report (filtered)", len(rows)],
        ["Numeric column", numeric_col or ""],
        ["Parsed count", (colstats or {}).get("parsed_count")],
        ["Missing count", (colstats or {}).get("missing_count")],
        ["Unparsable count", (colstats or {}).get("unparsable_count")],
        ["Parse ratio", f"{parse_ratio*100:.1f}%"],
    ]
    r = _write_kv(ws_a, kv1, title="Report summary")

    ws_a.cell(row=r, column=1, value="Numeric statistics").font = Font(bold=True, size=14)
    r += 2

    kv2 = [
        ["avg", st.get("avg")],
        ["median", st.get("median")],
        ["min", st.get("min")],
        ["max", st.get("max")],
        ["std", st.get("std")],
        ["q1", st.get("q1")],
        ["q3", st.get("q3")],
        ["iqr", st.get("iqr")],
    ]
    for k, v in kv2:
        ws_a.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_a.cell(row=r, column=2, value=v)
        r += 1

    r += 1
    ws_a.cell(row=r, column=1, value="Parsing note").font = Font(bold=True)
    ws_a.cell(row=r, column=2, value=(colstats or {}).get("parsing_note", ""))

    # 3) Charts sheet (data + charts)
    ws_c = wb.create_sheet(_safe_sheet_title("Charts"))

    # Histogram data table
    ws_c["A1"] = f"Histogram for: {numeric_col}"
    ws_c["A1"].font = Font(bold=True, size=12)

    ws_c.append(["bin", "count"])
    for lab, cnt in zip(hist_labels or [], hist_counts or []):
        ws_c.append([lab, int(cnt)])

    # Create histogram chart
    if hist_labels and hist_counts:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Histogram"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Bin"

        data_ref = Reference(ws_c, min_col=2, min_row=2, max_row=1 + len(hist_labels) + 1)
        cats_ref = Reference(ws_c, min_col=1, min_row=3, max_row=2 + len(hist_labels))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 22
        ws_c.add_chart(chart, "D2")

    # Top categories table
    start_row = (len(hist_labels) + 5) if hist_labels else 5
    ws_c[f"A{start_row}"] = f"Top values for: {cat_col}"
    ws_c[f"A{start_row}"].font = Font(bold=True, size=12)

    ws_c.append([])  # spacer
    ws_c.append(["value", "count"])
    for lab, cnt in zip(top_labels or [], top_counts or []):
        ws_c.append([lab, int(cnt)])

    if top_labels and top_counts:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Top values"
        chart2.y_axis.title = "Count"
        chart2.x_axis.title = "Value"

        # figure out table start for top
        # after spacer + header row:
        top_header_row = ws_c.max_row - len(top_labels) - 0  # header row already appended
        data_ref2 = Reference(ws_c, min_col=2, min_row=top_header_row, max_row=top_header_row + len(top_labels))
        cats_ref2 = Reference(ws_c, min_col=1, min_row=top_header_row + 1, max_row=top_header_row + len(top_labels))
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats_ref2)
        chart2.height = 10
        chart2.width = 22
        ws_c.add_chart(chart2, "D" + str(top_header_row))

    # widen columns on Charts
    ws_c.column_dimensions["A"].width = 40
    ws_c.column_dimensions["B"].width = 12

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()